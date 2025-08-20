from flask import Flask, request, abort
from linebot import LineBotApi, WebhookHandler
from linebot.exceptions import InvalidSignatureError
from linebot.models import MessageEvent, TextMessage, TextSendMessage, QuickReply, QuickReplyButton, MessageAction
from linebot.models import TemplateSendMessage, ButtonsTemplate, PostbackAction, PostbackEvent,FlexSendMessage

from openpyxl import load_workbook, Workbook
from oauth2client.service_account import ServiceAccountCredentials
import os,json
import gspread


app = Flask(__name__)



line_bot_api = LineBotApi('goBLuHUNvyvNlynFI+Xc/jgM/n8L/60A0X23kx+n5QM6Hp6IpDWyu5w6qP/hs6T6uawY6KX3Ijo915mGgPrDh3X35BUbBL8V7Fu55gJkA5/c86rN1hM6y4WrXlL/lLAQqNe+lK8BcAvQUfCW2WPSswdB04t89/1O/w1cDnyilFU=')
handler = WebhookHandler('0accce30a3ead3b05e956b69b96fda08')

SERVICE_ACCOUNT_INFO = os.environ.get('SERVICE_ACCOUNT_JSON')  # 下載的 JSON 金鑰

SHEET_ID = '1DfyNVTDO5RYe4AOUOAfgRY5ostiZqor_nbeYe-Igd7k'
SHEET_NAME = '工作表1'

scope = ['https://spreadsheets.google.com/feeds',
         'https://www.googleapis.com/auth/drive']
credentials = ServiceAccountCredentials.from_json_keyfile_dict(
    json.loads(SERVICE_ACCOUNT_INFO),
    scope
)

gc = gspread.authorize(credentials)
sheet = gc.open_by_key(SHEET_ID).worksheet(SHEET_NAME)

def save_user_to_sheet(user_id, display_name):
    all_user_ids = sheet.col_values(1)  # 假設 user_id 在第一欄
    if user_id not in all_user_ids:
        sheet.append_row([user_id, display_name])
        print(f"已新增使用者: {display_name} ({user_id})")
    else:
        print(f"使用者已存在: {display_name} ({user_id})")

@app.route("/")
def home():
    try:
        msg = request.args.get('msg')

        # 先讀取 Excel 裡的編號欄位
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # 找到標題「編號」是第幾欄
        headers = [cell.value for cell in ws[1]]
        if "編號" not in headers:
            raise ValueError("Excel 裡找不到「編號」欄位")
        id_col_index = headers.index("編號") + 1  # openpyxl 是 1-index

        # 讀取所有編號欄位資料（從第2列開始）
        user_ids = [row[id_col_index - 1].value for row in ws.iter_rows(min_row=2) if row[id_col_index - 1].value]

        if msg == '1':
            # 依序發送訊息給 Excel 裡的所有 user_id
            for user_id in user_ids:
                line_bot_api.push_message(user_id, TextSendMessage(text='你好'))
            return "已發送給 Excel 裡的所有編號"
        else:
            return 'ok'

    except Exception as e:
        print('error:', e)
        return '發送失敗.'


@app.route("/callback", methods=['POST'])
def callback():
    signature = request.headers.get('X-Line-Signature')
    body = request.get_data(as_text=True)

    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)

    return 'OK'

EXCEL_FILE = 'users.xlsx'

# # 如果 Excel 不存在，先建立一個
# if not os.path.exists(EXCEL_FILE):
#     wb = Workbook()
#     ws = wb.active
#     ws.append(["user_id", "display_name"])  # 標題欄
#     wb.save(EXCEL_FILE)

# def save_user_to_excel(user_id, display_name):
#     wb = load_workbook(EXCEL_FILE)
#     ws = wb.active

#     # 先檢查 user_id 是否已經存在
#     existing_ids = [row[0].value for row in ws.iter_rows(min_row=2)]
#     if user_id not in existing_ids:
#         ws.append([user_id, display_name])
#         wb.save(EXCEL_FILE)
#         print(f"已新增使用者: {display_name} ({user_id})")
#     else:
#         print(f"使用者已存在: {display_name} ({user_id})")


@handler.add(MessageEvent, message=TextMessage)
def handle_message(event):
    if event.source.type == "user":  # 只收集個人使用者
        user_id = event.source.user_id
        profile = line_bot_api.get_profile(user_id)
        display_name = profile.display_name
        
        print("使用者名字:", display_name)
        
        # 儲存到 Google Sheet
        save_user_to_sheet(user_id, display_name)

    user_msg = event.message.text
    print(f"收到訊息：{event.message.text}")
    if user_msg == "健保":
        buttons_template = TemplateSendMessage(
        alt_text='您想詢問哪種健保問題？',
        template=ButtonsTemplate(
            title='您想詢問哪種健保問題？',
            text='請選擇：',
            actions=[
                MessageAction(label='xin thẻ bảo hiểm', text='xin thẻ bảo hiểm y tế'),
                MessageAction(label='健保費用', text='健保費用'),
                MessageAction(label='健保卡忘記帶', text='健保卡忘記帶'),
                MessageAction(label='健保退費', text='健保退費'),
            ]
        )
    )
        line_bot_api.reply_message(event.reply_token, buttons_template)

    def button_with_line(label, text):
        return {
        "type": "box",
        "layout": "vertical",
        "contents": [
            {
                "type": "button",
                "style": "link",
                "action": {"type": "message", "label": label, "text": text},
                "color": "#0000FF"
            },
            {
                "type": "box",
                "layout": "vertical",
                "height": "1px",
                "backgroundColor": "#CCCCCC",
                "margin": "sm"
            }
        ]
    }

    if user_msg == "學習":
        flex_message = FlexSendMessage(
            alt_text="學業相關問題選單",
            contents={
                "type": "carousel",
                "contents": [
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🎓學籍與課程資訊",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("Tài khoản của trường", "Tài khoản của trường"),
                                button_with_line("Tra cứu thông tin học sinh", "Tra cứu thông tin học sinh"),
                                button_with_line("thời khóa biểu", "thời khóa biểu"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "📊學業與操行表現",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("Tra cứu kết quả học tập", "Tra cứu kết quả học tập"),
                                button_with_line("tình hình rèn luyện", "tình hình rèn luyện"),
                                button_with_line("ktra số buổi vắng học", "ktra số buổi vắng học"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "📝請假與課程調整",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("xin nghỉ phép", "xin nghỉ phép"),
                                button_with_line("các môn bị rớt", "các môn bị rớt"),
                                button_with_line("Học lại", "Học lại"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🎓 就學相關手續",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("các loại xin nghỉ", "các loại xin nghỉ"),
                                button_with_line("giấy xác nhận đang học", "giấy xác nhận đang học"),
                                button_with_line("Điều kiện tốt nghiệp", "Điều kiện tốt nghiệp"),
                            ]
                        }
                    }
                ]
            }
        )
        line_bot_api.reply_message(event.reply_token, flex_message)


    elif user_msg == "學校":
        flex_message = FlexSendMessage(
            alt_text="學校相關問題選單",
            contents={
                "type": "carousel",
                "contents": [
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🏫學校介紹與資訊",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("Video về trường", "Video về trường"),
                                button_with_line("trang web của trường", "trang web của trường"),
                                button_with_line("lịch", "lịch"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "💰 學雜費資訊",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("Thời gian đóng học phí", "Thời gian đóng học phí"),
                                button_with_line("tải đơn đóng học phí", "tải đơn đóng học phí"),
                                button_with_line("xin trả góp học phí", "xin trả góp học phí"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🏢 校內行政與系所位置",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("Vị trí các tòa nhà", "Vị trí các tòa nhà"),
                                button_with_line("Vị trí văn phòng các khoa", "Vị trí văn phòng các khoa"),
                                button_with_line("Vị trí quán ăn", "Vị trí quán ăn"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🍽 校園生活機能",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("Vị trí cây atm", "Vị trí cây atm"),
                                button_with_line("Vị trí tiệm photo", "Vị trí tiệm photo"),
                                button_with_line("Các địa điểm mua đồ", "Các địa điểm mua đồ"),
                            ]
                        }
                    }
                ]
            }
        )
        line_bot_api.reply_message(event.reply_token, flex_message)

    elif user_msg == "資料申請":
        flex_message = FlexSendMessage(
            alt_text="學校相關問題選單",
            contents={
                "type": "carousel",
                "contents": [
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "📘 TOCFL 準備指南",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("tạo tài khoản", "tạo tài khoản"),
                                button_with_line("THI TOCFL", "THI TOCFL"),
                                button_with_line("luyện thi thử", "luyện thi thử"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🏥 健保說明",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("xin BHYT", "xin bhyt"),
                                button_with_line("mất BHYT", "mất bhyt"),
                                button_with_line("Không xin BHYT", "Không xin bhyt"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🏦 開戶與辦證",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("Xin giấy phép lao động", "Xin giấy phép lao động"),
                                button_with_line("mở tài khoản của bưu điện", "mở tài khoản của bưu điện"),
                                button_with_line("quên tên đăng nhập", "quên tên đăng nhập"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🏦 證件展延",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("Gia hạn thẻ cư trú", "Gia hạn thẻ cư trú"),
                                button_with_line("Gia hạn giấy phép đi làm", "Gia hạn giấy phép đi làm"),
                            ]
                        }
                    }
                ]
            }
        )
        line_bot_api.reply_message(event.reply_token, flex_message)

    elif user_msg == "生活機能":
        flex_message = FlexSendMessage(
            alt_text="學校相關問題選單",
            contents={
                "type": "carousel",
                "contents": [
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🏥 健康醫療",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("hướng dẫn khám bệnh", "hướng dẫn khám bệnh"),
                                button_with_line("Thông tin khám bệnh", "Thông tin khám bệnh"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🚌 交通指引",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("Hướng dẫn đi xe buýt", "Hướng dẫn đi xe buýt"),
                                button_with_line("Hướng dẫn đi tàu điện ngầm", "Hướng dẫn đi tàu điện ngầm"),
                                button_with_line("xe đưa đón trong trường", "xe đưa đón trong trường"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🏡 生活打工",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("các App tìm việc", "các App tìm việc"),
                                button_with_line("App thuê nhà", "App thuê nhà"),
                                button_with_line("Quy định làm thêm", "Quy định làm thêm"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🏡 生活機能",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("Nạp tiền vào thẻ EasyCard", "Nạp tiền vào thẻ EasyCard"),
                                button_with_line("quy trình đăng ký YouBike", "quy trình đăng ký YouBike"),
                            ]
                        }
                    }
                ]
            }
        )
        line_bot_api.reply_message(event.reply_token, flex_message)

    elif user_msg == "機車相關業務":
        flex_message = FlexSendMessage(
            alt_text="學校相關問題選單",
            contents={
                "type": "carousel",
                "contents": [
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🛵 駕照與保險",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("thi bằng lái xe máy", "thi bằng lái xe máy"),
                                button_with_line("Tra cứu bảo hiểm xe máy", "Tra cứu bảo hiểm xe máy"),
                                button_with_line("Tra cứu phí nhiên liệu", "Tra cứu phí nhiên liệu"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "💸 交通費用與罰單",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("kiểm định xe", "kiểm định xe"),
                                button_with_line("kiểm tra giấy phạt", "kiểm tra giấy phạt"),
                                button_with_line("app xe may", "app xe may"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "📞 意外處理與電話指南",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("xảy ra tai nạn giao thông", "xảy ra tai nạn giao thông"),
                                button_with_line("Mục thường bị phạt", "Mục thường bị phạt"),
                                button_with_line("SỐ ĐIỆN THOẠI KHẨN CẤP", "SỐ ĐIỆN THOẠI KHẨN CẤP"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "📄 車輛查詢",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("đăng ký chính chủ", "đăng ký chính chủ"),
                                
                            ]
                        }
                    }
                ]
            }
        )
        line_bot_api.reply_message(event.reply_token, flex_message)

    elif user_msg == "其他項目":
        flex_message = FlexSendMessage(
            alt_text="學校相關問題選單",
            contents={
                "type": "carousel",
                "contents": [
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "📬 生活便利",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("lấy thư", "lấy thư"),
                                button_with_line("lấy hàng", "lấy hàng"),
                                button_with_line("mua sách", "mua sách"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🛏 住宿與外出",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("đăng ký về nước", "đăng ký về nước"),
                                button_with_line("Đơn xin đăng ký qua đêm", "Đơn xin đăng ký qua đêm"),
                                button_with_line("Hướng dẫn thủ tục dọn ra ngoài", "Hướng dẫn thủ tục dọn ra ngoài"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🎫 學生工具",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("đăng ký định danh thẻ học sinh", "đăng ký định danh thẻ học sinh"),
                                button_with_line("Nhặt được đồ thất lạc", "Nhặt được đồ thất lạc"),
                                button_with_line("sinh hoạt!!", "sinh hoạt!!"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🧩 知你所需",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("Khám phá con người thật", "Khám phá con người thật"),
                                button_with_line("ứng dụng phổ biến", "ứng dụng phổ biến"),
                                button_with_line("đăng ký số điện thoại di động", "đăng ký số điện thoại di động"),
                            ]
                        }
                    },
                    {
                        "type": "bubble",
                        "body": {
                            "type": "box",
                            "layout": "vertical",
                            "contents": [
                                {
                                    "type": "text",
                                    "text": "🧩 知你所需",
                                    "weight": "bold",
                                    "size": "xl",
                                    "align": "center"
                                },
                                button_with_line("Liên hệ với chúng tôi", "Liên hệ với chúng tôi"),
                                button_with_line("Mất thẻ sinh viên", "Mất thẻ sinh viên"),                    
                            ]
                        }
                    }
                ]
            }
        )
        line_bot_api.reply_message(event.reply_token, flex_message) 


        flex_message = FlexSendMessage(
    alt_text='您想詢問哪種課程問題？',
    contents={
        "type": "bubble",
        "body": {
            "type": "box",
            "layout": "vertical",
            "contents": [
                {
                    "type": "text",
                    "text": "請選擇課程相關問題",
                    "weight": "bold",
                    "size": "xl",
                    "align": "center",
                    "marginBottom": "lg"
                },
                {
                    "type": "box",
                    "layout": "horizontal",
                    "spacing": "sm",
                    "contents": [
                        {
                            "type": "box",
                            "layout": "vertical",
                            "flex": 1,
                            "cornerRadius": "md",
                            "paddingAll": "5px",
                            "contents": [
                                {
                                    "type": "button",
                                    "style": "link",
                                    "action": {
                                        "type": "message",
                                        "label": "課程報名",
                                        "text": "課程報名"
                                    },
                                    "color": "#4848FF",
                                    "height": "sm"
                                },
                                {
                                    "type": "separator",
                                    "margin": "sm",
                                    "color": "#DDDDDD"
                                },
                                {
                                    "type": "button",
                                    "style": "link",
                                    "action": {
                                        "type": "message",
                                        "label": "課程費用",
                                        "text": "課程費用"
                                    },
                                    "color": "#0000FF",
                                    "height": "sm"
                                },
                                {
                                    "type": "separator",
                                    "margin": "sm",
                                    "color": "#DDDDDD"
                                },
                                {
                                    "type": "button",
                                    "style": "link",
                                    "action": {
                                        "type": "message",
                                        "label": "課程證明",
                                        "text": "課程證明"
                                    },
                                    "color": "#0000FF",
                                    "height": "sm"
                                }
                            ]
                        },
                        {
                            "type": "box",
                            "layout": "vertical",
                            "flex": 1,
                            "cornerRadius": "md",
                            "paddingAll": "5px",
                            "contents": [
                                {
                                    "type": "button",
                                    "style": "link",
                                    "action": {
                                        "type": "message",
                                        "label": "上課地點",
                                        "text": "上課地點"
                                    },
                                    "color": "#0000FF",
                                    "height": "sm"
                                },
                                {
                                    "type": "separator",
                                    "margin": "sm",
                                    "color": "#DDDDDD"
                                },
                                {
                                    "type": "button",
                                    "style": "link",
                                    "action": {
                                        "type": "message",
                                        "label": "上課時間",
                                        "text": "上課時間"
                                    },
                                    "color": "#0000FF",
                                    "height": "sm"
                                },
                                {
                                    "type": "separator",
                                    "margin": "sm",
                                    "color": "#DDDDDD"
                                },
                                {
                                    "type": "button",
                                    "style": "link",
                                    "action": {
                                        "type": "message",
                                        "label": "其他問題",
                                        "text": "其他問題"
                                    },
                                    "color": "#0000FF",
                                    "height": "sm"
                                }
                            ]
                        }
                    ]
                }
            ]
        }
    }
)

        line_bot_api.reply_message(event.reply_token, flex_message)

    
    


    
# 處理使用者點擊 postback 按鈕
@handler.add(PostbackEvent)
def handle_postback(event):
    data = event.postback.data

    if data == '加入健保':
        reply_text = "加入健保"
    elif data == '健保費用':
        reply_text = "健保費用"
    elif data == '健保卡忘記帶':
        reply_text = "健保卡忘記帶"
    elif data == '健保退費':
        reply_text = "健保退費"
    else:
        reply_text = "您選擇了：" + data

    line_bot_api.reply_message(
        event.reply_token,
        TextSendMessage(text=reply_text)
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)